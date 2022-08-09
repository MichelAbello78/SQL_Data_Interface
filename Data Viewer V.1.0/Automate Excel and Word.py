from email.mime import image
from multiprocessing import current_process
import openpyxl as xl
from openpyxl.chart import Linechart, Reference
import win32com.client
import PIL
import os
import sys
from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu
import random
import datetime
import matplotlib.pyplot as plt

workbook= xl.load_workbook('Libro1.xlsx')
sheet= workbook['sheet1']

for row in range (2, sheet.max_column + 1):
    actual= sheet.cell(row, 2)
    voltage=sheet.cell(row, 3)
    power=float(actual.value)*float(voltage.value)
    power_cell=sheet.cell(row, 1)
    power_cell.value=power

values=Reference(sheet, min_row=2, max_row= sheet.max_row, min_col=1, max_col=1)
chart=Linechart()
chart.y_axist.title='Power'
chart.x_axits.title='Index'
chart.add_data(values)
sheet.add_chart(chart, 'e2')
workbook.save('Libro1.xlsx')
input_file="D:/Users/Michel Abello/Downloads/Libro1.xlsx"
output_image="D:/Users/Michel Abello/Downloads/chart.png"
operation=win32com.client.Dispatch("Excel.Application")
operation.Visible=0
operation.DisplayAlerts=0
workbook2=operation.Workbooks.Open(input_file)
sheet2=operation.Sheets(1)

for x, chart in enumerate(sheet2.Shapes):
    chart.Copy()
    image=ImageGrab.grabclipboard()
    image.save(output_image, 'png')

    pass
